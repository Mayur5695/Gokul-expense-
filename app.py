import streamlit as st
import pandas as pd
import os
from datetime import datetime, date
import io

# ─── Page Config ────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Expense Tracker Portal",
    page_icon="💰",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ─── Custom CSS ──────────────────────────────────────────────────────────────
st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Rajdhani:wght@400;600;700&family=Noto+Sans+Devanagari:wght@400;500;600&display=swap');

    html, body, [class*="css"] {
        font-family: 'Rajdhani', 'Noto Sans Devanagari', sans-serif;
    }

    .stApp {
        background: linear-gradient(135deg, #0f0c29, #302b63, #24243e);
        min-height: 100vh;
    }

    /* Title */
    .main-title {
        font-family: 'Rajdhani', sans-serif;
        font-size: 2.8rem;
        font-weight: 700;
        text-align: center;
        background: linear-gradient(90deg, #f7971e, #ffd200);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        padding: 0.5rem 0 0.2rem 0;
        letter-spacing: 2px;
    }

    .subtitle {
        text-align: center;
        color: #a0aec0;
        font-size: 1rem;
        margin-bottom: 1.5rem;
    }

    /* Cards */
    .stat-card {
        background: rgba(255,255,255,0.05);
        border: 1px solid rgba(247,151,30,0.3);
        border-radius: 16px;
        padding: 1.2rem 1.5rem;
        text-align: center;
        backdrop-filter: blur(10px);
    }
    .stat-card h3 {
        color: #ffd200;
        font-size: 1.8rem;
        font-weight: 700;
        margin: 0;
    }
    .stat-card p {
        color: #a0aec0;
        font-size: 0.85rem;
        margin: 0.2rem 0 0 0;
    }

    /* Person badge */
    .person-badge {
        display: inline-block;
        padding: 0.25rem 0.8rem;
        border-radius: 20px;
        font-size: 0.8rem;
        font-weight: 600;
        letter-spacing: 1px;
    }
    .badge-suhail { background: rgba(99,179,237,0.2); color: #63b3ed; border: 1px solid #63b3ed; }
    .badge-mayur  { background: rgba(154,230,180,0.2); color: #9ae6b4; border: 1px solid #9ae6b4; }
    .badge-rahul  { background: rgba(252,182,159,0.2); color: #fca18f; border: 1px solid #fca18f; }

    /* Sidebar */
    section[data-testid="stSidebar"] {
        background: rgba(15,12,41,0.95);
        border-right: 1px solid rgba(247,151,30,0.2);
    }

    /* Buttons */
    .stButton > button {
        background: linear-gradient(90deg, #f7971e, #ffd200);
        color: #1a1a2e;
        font-weight: 700;
        border: none;
        border-radius: 10px;
        padding: 0.5rem 1.5rem;
        font-family: 'Rajdhani', sans-serif;
        font-size: 1rem;
        letter-spacing: 1px;
        transition: all 0.3s;
    }
    .stButton > button:hover {
        transform: translateY(-2px);
        box-shadow: 0 8px 20px rgba(247,151,30,0.4);
    }

    /* Delete button */
    .delete-btn > button {
        background: linear-gradient(90deg, #e53e3e, #c53030) !important;
        color: white !important;
        font-size: 0.8rem !important;
        padding: 0.3rem 0.8rem !important;
    }

    /* Input fields */
    .stTextInput > div > div > input,
    .stNumberInput > div > div > input,
    .stSelectbox > div > div,
    .stTextArea > div > div > textarea {
        background: rgba(255,255,255,0.08) !important;
        border: 1px solid rgba(247,151,30,0.3) !important;
        border-radius: 8px !important;
        color: white !important;
    }

    .stDateInput > div > div > input {
        background: rgba(255,255,255,0.08) !important;
        border: 1px solid rgba(247,151,30,0.3) !important;
        color: white !important;
    }

    /* Dataframe */
    .stDataeditor {
        border-radius: 12px;
        overflow: hidden;
    }

    /* Divider */
    hr { border-color: rgba(247,151,30,0.2); }

    /* Section headers */
    .section-header {
        font-size: 1.3rem;
        font-weight: 700;
        color: #ffd200;
        border-left: 4px solid #f7971e;
        padding-left: 0.8rem;
        margin: 1rem 0 0.8rem 0;
    }

    /* Success/Error messages */
    .stSuccess { background: rgba(72,187,120,0.15) !important; border-radius: 10px !important; }
    .stError   { background: rgba(229,62,62,0.15)  !important; border-radius: 10px !important; }
    .stWarning { background: rgba(247,151,30,0.15) !important; border-radius: 10px !important; }
</style>
""", unsafe_allow_html=True)

# ─── Data File ───────────────────────────────────────────────────────────────
DATA_FILE = "expenses.csv"

PERSONS = ["Suhail", "Mayur", "Rahul"]

EXPENSE_CATEGORIES = [
    "Warehouse Rent",
    "Warehouse Rent Advance",
    "Vehicle Expense",
    "Vehicle Expense Advance",
    "Petrol",
    "Petrol Advance",
    "Salary",
    "Office Supplies",
    "Electricity Bill",
    "Internet/Phone Bill",
    "Food & Meals",
    "Travel Expense",
    "Maintenance & Repair",
    "Miscellaneous",
    "Other"
]

TRANSACTION_TYPES = ["Expense", "Income"]
ACCOUNT_NAMES = ["Cash", "Bank", "Credit Card", "Accounts Payable", "Accounts Receivable", "Capital"]
COLUMNS = ["ID", "Date", "Person", "Account", "Type", "Category", "Amount", "Note"]

# ─── Helper Functions ─────────────────────────────────────────────────────────
def load_data():
    if os.path.exists(DATA_FILE):
        df = pd.read_csv(DATA_FILE)
        if df.empty:
            df = pd.DataFrame(columns=COLUMNS)
        else:
            for col in COLUMNS:
                if col not in df.columns:
                    if col == "Account":
                        df[col] = "Cash"
                    elif col == "Type":
                        df[col] = "Expense"
                    else:
                        df[col] = "" if col in ["Person", "Category", "Note"] else 0
            df = df[COLUMNS]
        df["Date"] = pd.to_datetime(df["Date"], errors="coerce")
        df["Amount"] = pd.to_numeric(df["Amount"], errors="coerce").fillna(0)
        return df
    return pd.DataFrame(columns=COLUMNS)


def save_data(df):
    df.to_csv(DATA_FILE, index=False)


def next_id(df):
    if df.empty or "ID" not in df.columns:
        return 1
    return int(df["ID"].max()) + 1


def format_inr(amount):
    return f"₹{amount:,.2f}"


def badge_html(person):
    cls = f"badge-{person.lower()}"
    return f'<span class="person-badge {cls}">{person}</span>'


# ─── Load Data ────────────────────────────────────────────────────────────────
df = load_data()

# ─── Sidebar Navigation ───────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("### � Accounting Portal")
    st.markdown("---")
    nav = st.radio(
        "Navigation",
        ["📊 Dashboard", "➕ Add Transaction", "📋 Ledger", "📅 Accounting Summary", "👤 Add Person"],
        label_visibility="collapsed"
    )
    st.markdown("---")
    
    # Quick stats
    if not df.empty:
        st.markdown("**Quick Stats**")
        signed_total = (df["Amount"] * df["Type"].map({"Expense": -1, "Income": 1}).fillna(1)).sum()
        st.markdown(f"🔢 Total Transactions: **{len(df)}**")
        st.markdown(f"💵 Net Amount: **{format_inr(signed_total)}**")
        
        this_month = df[df["Date"].dt.month == datetime.now().month].copy()
        signed_month = (this_month["Amount"] * this_month["Type"].map({"Expense": -1, "Income": 1}).fillna(1)).sum()
        st.markdown(f"📆 This Month: **{format_inr(signed_month)}**")

# ─── Load persons from session state ─────────────────────────────────────────
if "persons" not in st.session_state:
    st.session_state.persons = PERSONS.copy()

# ─── Main Title ───────────────────────────────────────────────────────────────
st.markdown('<div class="main-title">💼 ACCOUNTING PORTAL</div>', unsafe_allow_html=True)
st.markdown('<div class="subtitle">Suhail • Mayur • Rahul — Business Accounting Software</div>', unsafe_allow_html=True)
st.markdown("---")

# ════════════════════════════════════════════════════════════════════════════════
# DASHBOARD
# ════════════════════════════════════════════════════════════════════════════════
if nav == "📊 Dashboard":
    df = load_data()

    if df.empty:
        st.info("📭 कोणतीही entries नाहीत. 'Add Transaction' वर जाऊन पहिली entry add करा!")
    else:
        # Top stats
        col1, col2, col3, col4 = st.columns(4)
        signed_amounts = df["Amount"] * df["Type"].map({"Expense": -1, "Income": 1}).fillna(1)
        total_all = signed_amounts.sum()
        this_m = df[df["Date"].dt.month == datetime.now().month]
        total_m = (this_m["Amount"] * this_m["Type"].map({"Expense": -1, "Income": 1}).fillna(1)).sum()
        total_entries = len(df)

        with col1:
            st.markdown(f'<div class="stat-card"><h3>{format_inr(total_all)}</h3><p>Total All Time</p></div>', unsafe_allow_html=True)
        with col2:
            st.markdown(f'<div class="stat-card"><h3>{format_inr(total_m)}</h3><p>This Month</p></div>', unsafe_allow_html=True)
        with col3:
            st.markdown(f'<div class="stat-card"><h3>{total_entries}</h3><p>Total Entries</p></div>', unsafe_allow_html=True)
        with col4:
            avg = total_all / total_entries if total_entries else 0
            st.markdown(f'<div class="stat-card"><h3>{format_inr(avg)}</h3><p>Avg per Entry</p></div>', unsafe_allow_html=True)

        st.markdown("")

        # Per Person Summary
        st.markdown('<div class="section-header">👥 Person-wise Summary</div>', unsafe_allow_html=True)
        cols = st.columns(len(st.session_state.persons))
        for i, person in enumerate(st.session_state.persons):
            person_data = df[df["Person"] == person]
            person_total = person_data["Amount"].sum()
            person_count = len(person_data)
            with cols[i]:
                st.markdown(f"""
                <div class="stat-card">
                    <h3>{format_inr(person_total)}</h3>
                    <p>{person} — {person_count} entries</p>
                </div>
                """, unsafe_allow_html=True)

        st.markdown("")

        # Category breakdown
        st.markdown('<div class="section-header">📂 Category-wise Total</div>', unsafe_allow_html=True)
        cat_summary = df.groupby("Category")["Amount"].sum().reset_index()
        cat_summary = cat_summary.sort_values("Amount", ascending=False)
        cat_summary["Amount"] = cat_summary["Amount"].apply(format_inr)
        cat_summary.columns = ["Category", "Total Amount"]
        st.dataframe(cat_summary, use_container_width=True, hide_index=True)

        # Recent transactions
        st.markdown('<div class="section-header">🕐 Recent Transactions (Last 10)</div>', unsafe_allow_html=True)
        recent = df.sort_values("Date", ascending=False).head(10).copy()
        recent["Date"] = recent["Date"].dt.strftime("%d %b %Y")
        recent["Amount"] = recent["Amount"].apply(format_inr)
        st.dataframe(recent[["ID", "Date", "Person", "Account", "Type", "Category", "Amount", "Note"]], use_container_width=True, hide_index=True)

        # Account balances
        st.markdown('<div class="section-header">🏦 Account Balances</div>', unsafe_allow_html=True)
        account_summary = df.groupby("Account")["Amount"].sum().reset_index()
        account_summary["Amount"] = account_summary["Amount"].apply(format_inr)
        account_summary.columns = ["Account", "Balance"]
        st.dataframe(account_summary, use_container_width=True, hide_index=True)

        # Recent transaction edit section
        with st.expander('✏️ Recent Transaction Edit करा', expanded=True):
            st.info('Recent transactions पैकी एक transaction ID निवडा आणि fields बदला.')
            recent_ids = recent["ID"].astype(int).tolist()
            if recent_ids:
                selected_id = st.selectbox(
                    "Edit करण्यासाठी Recent Transaction ID निवडा",
                    recent_ids,
                    format_func=lambda x: f"ID {x} — {df[df['ID']==x]['Person'].values[0]} | {df[df['ID']==x]['Account'].values[0]} | {df[df['ID']==x]['Type'].values[0]} | ₹{df[df['ID']==x]['Amount'].values[0]:,.2f} | {df[df['ID']==x]['Date'].dt.strftime('%d %b %Y').values[0]}"
                )
                selected_entry = df[df["ID"] == selected_id].iloc[0]
                with st.form("recent_entry_edit_form"):
                    col1, col2 = st.columns(2)
                    with col1:
                        edit_person = st.selectbox(
                            "👤 Person",
                            st.session_state.persons,
                            index=st.session_state.persons.index(selected_entry["Person"]) if selected_entry["Person"] in st.session_state.persons else 0
                        )
                        edit_account = st.selectbox(
                            "🏦 Account",
                            ACCOUNT_NAMES,
                            index=ACCOUNT_NAMES.index(selected_entry["Account"]) if selected_entry["Account"] in ACCOUNT_NAMES else 0
                        )
                        edit_type = st.selectbox(
                            "📌 Type",
                            TRANSACTION_TYPES,
                            index=TRANSACTION_TYPES.index(selected_entry["Type"]) if selected_entry["Type"] in TRANSACTION_TYPES else 0
                        )
                        edit_date = st.date_input("📅 Date", value=selected_entry["Date"].date())
                        edit_category = st.selectbox(
                            "🏷️ Category",
                            EXPENSE_CATEGORIES,
                            index=EXPENSE_CATEGORIES.index(selected_entry["Category"]) if selected_entry["Category"] in EXPENSE_CATEGORIES else 0
                        )
                    with col2:
                        edit_amount = st.number_input(
                            "💵 Amount ₹",
                            min_value=0.0,
                            step=100.0,
                            format="%.2f",
                            value=float(selected_entry["Amount"])
                        )
                        edit_note = st.text_area("📝 Note", value=selected_entry["Note"], height=120)

                    edit_submitted = st.form_submit_button("✅ Save Recent Transaction Changes", use_container_width=True)
                    if edit_submitted:
                        if edit_amount <= 0:
                            st.error("⚠️ Amount 0 पेक्षा जास्त असणे आवश्यक आहे!")
                        else:
                            df_edit = load_data()
                            idx = df_edit.index[df_edit["ID"] == selected_id]
                            if not idx.empty:
                                idx = idx[0]
                                df_edit.at[idx, "Person"] = edit_person
                                df_edit.at[idx, "Account"] = edit_account
                                df_edit.at[idx, "Type"] = edit_type
                                df_edit.at[idx, "Date"] = pd.Timestamp(edit_date)
                                df_edit.at[idx, "Category"] = edit_category
                                df_edit.at[idx, "Amount"] = edit_amount
                                df_edit.at[idx, "Note"] = edit_note.strip()
                                save_data(df_edit)
                                st.success(f"✅ Recent transaction ID {selected_id} updated successfully!")
                                st.rerun()
            else:
                st.info("Recent transactions उपलब्ध नाहीत.")


# ════════════════════════════════════════════════════════════════════════════════
# ADD TRANSACTION
# ════════════════════════════════════════════════════════════════════════════════
elif nav == "➕ Add Transaction":
    st.markdown('<div class="section-header">➕ नवीन Transaction Add करा</div>', unsafe_allow_html=True)

    with st.form("add_expense_form", clear_on_submit=True):
        col1, col2 = st.columns(2)

        with col1:
            person = st.selectbox("👤 व्यक्तीचे नाव (Person)", st.session_state.persons)
            transaction_date = st.date_input("📅 तारीख (Date)", value=date.today())
            transaction_type = st.selectbox("📌 Transaction Type", TRANSACTION_TYPES)
            account = st.selectbox("🏦 Account", ACCOUNT_NAMES)

        with col2:
            category = st.selectbox("🏷️ Category", EXPENSE_CATEGORIES)
            amount = st.number_input("💵 रक्कम (Amount ₹)", min_value=0.0, step=100.0, format="%.2f")
            note = st.text_area("📝 नोट (Note)", placeholder="Optional - additional details...", height=120)

        st.markdown("")
        submitted = st.form_submit_button("✅ Transaction Save करा", use_container_width=True)

        if submitted:
            if amount <= 0:
                st.error("⚠️ Amount 0 पेक्षा जास्त असणे आवश्यक आहे!")
            else:
                df = load_data()
                new_row = pd.DataFrame([{
                    "ID": next_id(df),
                    "Date": pd.Timestamp(transaction_date),
                    "Person": person,
                    "Account": account,
                    "Type": transaction_type,
                    "Category": category,
                    "Amount": amount,
                    "Note": note.strip() if note else ""
                }])
                df = pd.concat([df, new_row], ignore_index=True)
                save_data(df)
                st.success(f"✅ **{person}** चा ₹{amount:,.2f} चा **{transaction_type}** transaction successfully saved!")
                st.balloons()


# ════════════════════════════════════════════════════════════════════════════════
# LEDGER
# ════════════════════════════════════════════════════════════════════════════════
elif nav == "📋 Ledger":
    st.markdown('<div class="section-header">📋 Ledger Transactions</div>', unsafe_allow_html=True)
    df = load_data()

    if df.empty:
        st.info("📭 कोणतेही records नाहीत.")
    else:
        # Filters
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            filter_person = st.multiselect("👤 Person Filter", ["All"] + st.session_state.persons, default=["All"])
        with col2:
            filter_cat = st.multiselect("🏷️ Category Filter", ["All"] + EXPENSE_CATEGORIES, default=["All"])
        with col3:
            filter_account = st.multiselect("🏦 Account Filter", ["All"] + ACCOUNT_NAMES, default=["All"])
        with col4:
            filter_type = st.multiselect("📌 Type Filter", ["All"] + TRANSACTION_TYPES, default=["All"])

        filter_month = st.selectbox("📅 Month Filter", ["All"] + [
            f"{m:02d}/{y}" for y in sorted(df["Date"].dt.year.unique(), reverse=True)
            for m in range(1, 13)
            if not df[(df["Date"].dt.month == m) & (df["Date"].dt.year == y)].empty
        ])

        # Apply filters
        filtered = df.copy()
        if "All" not in filter_person and filter_person:
            filtered = filtered[filtered["Person"].isin(filter_person)]
        if "All" not in filter_cat and filter_cat:
            filtered = filtered[filtered["Category"].isin(filter_cat)]
        if "All" not in filter_account and filter_account:
            filtered = filtered[filtered["Account"].isin(filter_account)]
        if "All" not in filter_type and filter_type:
            filtered = filtered[filtered["Type"].isin(filter_type)]
        if filter_month != "All":
            m, y = int(filter_month.split("/")[0]), int(filter_month.split("/")[1])
            filtered = filtered[(filtered["Date"].dt.month == m) & (filtered["Date"].dt.year == y)]

        filtered = filtered.sort_values("Date", ascending=False)

        # Summary of filtered
        signed_filtered_total = (filtered["Amount"] * filtered["Type"].map({"Expense": -1, "Income": 1}).fillna(1)).sum()
        st.markdown(f"**{len(filtered)} entries found — Net Total: {format_inr(signed_filtered_total)}**")
        st.markdown("")

        # Edit section
        st.markdown("---")
        with st.expander('✏️ Entry Edit करा', expanded=True):
            st.info("खालील ID निवडा आणि नंतर त्या entry चे तपशील edit करा.")

            if not filtered.empty:
                id_options = filtered["ID"].astype(int).tolist()
                edit_id = st.selectbox(
                    "Edit करण्यासाठी Entry ID निवडा",
                    id_options,
                    format_func=lambda x: f"ID {x} — {filtered[filtered['ID']==x]['Person'].values[0]} | {filtered[filtered['ID']==x]['Category'].values[0]} | ₹{filtered[filtered['ID']==x]['Amount'].values[0]:,.2f} | {filtered[filtered['ID']==x]['Date'].dt.strftime('%d %b %Y').values[0]}"
                )

                edit_entry = filtered[filtered["ID"] == edit_id].iloc[0]
                with st.form("edit_expense_form"):
                    col1, col2 = st.columns(2)
                    with col1:
                        edit_person = st.selectbox(
                            "👤 Person",
                            st.session_state.persons,
                            index=st.session_state.persons.index(edit_entry["Person"]) if edit_entry["Person"] in st.session_state.persons else 0
                        )
                        edit_account = st.selectbox(
                            "🏦 Account",
                            ACCOUNT_NAMES,
                            index=ACCOUNT_NAMES.index(edit_entry["Account"]) if edit_entry["Account"] in ACCOUNT_NAMES else 0
                        )
                        edit_type = st.selectbox(
                            "📌 Type",
                            TRANSACTION_TYPES,
                            index=TRANSACTION_TYPES.index(edit_entry["Type"]) if edit_entry["Type"] in TRANSACTION_TYPES else 0
                        )
                        edit_date = st.date_input("📅 Date", value=edit_entry["Date"].date())
                        edit_category = st.selectbox(
                            "🏷️ Category",
                            EXPENSE_CATEGORIES,
                            index=EXPENSE_CATEGORIES.index(edit_entry["Category"]) if edit_entry["Category"] in EXPENSE_CATEGORIES else 0
                        )

                    with col2:
                        edit_amount = st.number_input(
                            "💵 Amount ₹",
                            min_value=0.0,
                            step=100.0,
                            format="%.2f",
                            value=float(edit_entry["Amount"])
                        )
                        edit_note = st.text_area("📝 Note", value=edit_entry["Note"], height=120)

                    edit_submitted = st.form_submit_button("✅ Save Changes", use_container_width=True)
                    if edit_submitted:
                        if edit_amount <= 0:
                            st.error("⚠️ Amount 0 पेक्षा जास्त असणे आवश्यक आहे!")
                        else:
                            df = load_data()
                            idx = df.index[df["ID"] == edit_id]
                            if not idx.empty:
                                idx = idx[0]
                                df.at[idx, "Person"] = edit_person
                                df.at[idx, "Account"] = edit_account
                                df.at[idx, "Type"] = edit_type
                                df.at[idx, "Date"] = pd.Timestamp(edit_date)
                                df.at[idx, "Category"] = edit_category
                                df.at[idx, "Amount"] = edit_amount
                                df.at[idx, "Note"] = edit_note.strip()
                                save_data(df)
                                st.success(f"✅ Entry ID {edit_id} successfully updated!")
                                st.rerun()

        # Display with delete option
        display_df = filtered.copy()
        display_df["Date_Display"] = display_df["Date"].dt.strftime("%d %b %Y")
        display_df["Amount_Display"] = display_df["Amount"].apply(lambda x: f"₹{x:,.2f}")

        # Show table
        show_cols = ["ID", "Date_Display", "Person", "Account", "Type", "Category", "Amount_Display", "Note"]
        renamed = display_df[show_cols].rename(columns={
            "Date_Display": "Date",
            "Amount_Display": "Amount"
        })
        st.dataframe(renamed, use_container_width=True, hide_index=True)

        # Delete section
        st.markdown("---")
        st.markdown('<div class="section-header">🗑️ Entry Delete करा</div>', unsafe_allow_html=True)

        if not filtered.empty:
            id_options = filtered["ID"].astype(int).tolist()
            del_id = st.selectbox(
                "Delete करायची Entry ID निवडा",
                id_options,
                format_func=lambda x: f"ID {x} — {filtered[filtered['ID']==x]['Person'].values[0]} | {filtered[filtered['ID']==x]['Account'].values[0]} | {filtered[filtered['ID']==x]['Type'].values[0]} | {filtered[filtered['ID']==x]['Category'].values[0]} | ₹{filtered[filtered['ID']==x]['Amount'].values[0]:,.2f} | {filtered[filtered['ID']==x]['Date'].dt.strftime('%d %b %Y').values[0]}"
            )

            col_del, col_cancel = st.columns([1, 3])
            with col_del:
                if st.button("🗑️ Delete Entry", type="primary"):
                    df = load_data()
                    df = df[df["ID"] != del_id]
                    save_data(df)
                    st.success(f"✅ Entry ID {del_id} successfully deleted!")
                    st.rerun()


# ════════════════════════════════════════════════════════════════════════════════
# ACCOUNTING SUMMARY + EXCEL DOWNLOAD
# ════════════════════════════════════════════════════════════════════════════════
elif nav == "📅 Accounting Summary":
    st.markdown('<div class="section-header">📅 Accounting Summary & Excel Download</div>', unsafe_allow_html=True)
    df = load_data()

    if df.empty:
        st.info("📭 कोणतेही records नाहीत.")
    else:
        df["Month"] = df["Date"].dt.to_period("M")
        df["Month_Str"] = df["Date"].dt.strftime("%B %Y")

        # Month selector
        months_available = df["Month_Str"].unique().tolist()
        selected_month = st.selectbox("📅 Month निवडा", ["All Months"] + months_available)

        if selected_month == "All Months":
            filtered = df.copy()
            title_str = "All Months"
        else:
            filtered = df[df["Month_Str"] == selected_month].copy()
            title_str = selected_month

        signed_summary_total = (filtered["Amount"] * filtered["Type"].map({"Expense": -1, "Income": 1}).fillna(1)).sum()
        st.markdown(f"### 📊 Summary — {title_str}")
        st.markdown(f"**Net Total: {format_inr(signed_summary_total)} | {len(filtered)} Transactions**")
        st.markdown("")

        # Person-wise for selected month
        col1, col2 = st.columns(2)
        with col1:
            st.markdown("**Person-wise Breakdown**")
            person_summary = filtered.groupby("Person")["Amount"].sum().reset_index()
            person_summary["Amount"] = person_summary["Amount"].apply(format_inr)
            person_summary.columns = ["Person", "Total Amount"]
            st.dataframe(person_summary, use_container_width=True, hide_index=True)

        with col2:
            st.markdown("**Category-wise Breakdown**")
            cat_summary = filtered.groupby("Category")["Amount"].sum().reset_index()
            cat_summary = cat_summary.sort_values("Amount", ascending=False)
            cat_summary["Amount"] = cat_summary["Amount"].apply(format_inr)
            cat_summary.columns = ["Category", "Total Amount"]
            st.dataframe(cat_summary, use_container_width=True, hide_index=True)

        col3, col4 = st.columns(2)
        with col3:
            st.markdown("**Transaction Type Breakdown**")
            type_summary = filtered.groupby("Type")["Amount"].sum().reset_index()
            type_summary["Amount"] = type_summary["Amount"].apply(format_inr)
            type_summary.columns = ["Type", "Total Amount"]
            st.dataframe(type_summary, use_container_width=True, hide_index=True)

        with col4:
            st.markdown("**Account-wise Breakdown**")
            acct_summary = filtered.groupby("Account")["Amount"].sum().reset_index()
            acct_summary["Amount"] = acct_summary["Amount"].apply(format_inr)
            acct_summary.columns = ["Account", "Total Amount"]
            st.dataframe(acct_summary, use_container_width=True, hide_index=True)

        st.markdown("---")

        # Excel Export
        st.markdown('<div class="section-header">⬇️ Excel मध्ये Download करा</div>', unsafe_allow_html=True)

        export_df = filtered.copy()
        export_df["Date"] = export_df["Date"].dt.strftime("%d-%m-%Y")
        export_df = export_df.drop(columns=["Month", "Month_Str"], errors="ignore")
        export_df = export_df.sort_values("Date", ascending=False)

        # Build Excel in memory
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            # Main data sheet
            export_df.to_excel(writer, sheet_name="Expenses", index=False)
            ws = writer.sheets["Expenses"]

            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            # Header styling
            header_fill = PatternFill("solid", fgColor="F7971E")
            header_font = Font(bold=True, color="1A1A2E", size=12)
            for col_num, cell in enumerate(ws[1], 1):
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            # Column widths
            col_widths = {"ID": 6, "Date": 14, "Person": 12, "Category": 28, "Amount": 14, "Note": 35}
            for col_num, column in enumerate(ws.columns, 1):
                col_name = ws.cell(1, col_num).value
                ws.column_dimensions[get_column_letter(col_num)].width = col_widths.get(col_name, 15)

            # Alternating rows
            light_fill = PatternFill("solid", fgColor="F0EFF8")
            for row_idx, row in enumerate(ws.iter_rows(min_row=2), 2):
                if row_idx % 2 == 0:
                    for cell in row:
                        cell.fill = light_fill

            # Summary sheet
            summary_data = []
            summary_data.append(["EXPENSE SUMMARY — " + title_str, ""])
            summary_data.append(["", ""])
            summary_data.append(["Total Amount", filtered["Amount"].sum()])
            summary_data.append(["Total Entries", len(filtered)])
            summary_data.append(["", ""])
            summary_data.append(["PERSON-WISE BREAKDOWN", ""])
            for _, row in filtered.groupby("Person")["Amount"].sum().reset_index().iterrows():
                summary_data.append([row["Person"], row["Amount"]])
            summary_data.append(["", ""])
            summary_data.append(["CATEGORY-WISE BREAKDOWN", ""])
            for _, row in filtered.groupby("Category")["Amount"].sum().reset_index().sort_values("Amount", ascending=False).iterrows():
                summary_data.append([row["Category"], row["Amount"]])

            summary_df = pd.DataFrame(summary_data, columns=["Label", "Value"])
            summary_df.to_excel(writer, sheet_name="Summary", index=False)

            ws2 = writer.sheets["Summary"]
            ws2.column_dimensions["A"].width = 35
            ws2.column_dimensions["B"].width = 18
            for cell in ws2["A"]:
                if cell.value and str(cell.value).isupper() and cell.value.strip():
                    cell.font = Font(bold=True, color="F7971E", size=11)

        output.seek(0)
        filename = f"Expenses_{title_str.replace(' ', '_')}.xlsx"
        st.download_button(
            label="⬇️ Excel Download करा",
            data=output,
            file_name=filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
        st.success(f"✅ '{filename}' तयार आहे — वर click करून download करा!")


# ════════════════════════════════════════════════════════════════════════════════
# ADD PERSON
# ════════════════════════════════════════════════════════════════════════════════
elif nav == "👤 Add Person":
    st.markdown('<div class="section-header">👤 नवीन व्यक्ती Add करा</div>', unsafe_allow_html=True)

    st.markdown("**सध्याचे लोक (Current Persons):**")
    cols = st.columns(len(st.session_state.persons))
    for i, p in enumerate(st.session_state.persons):
        cols[i].markdown(f"👤 **{p}**")

    st.markdown("---")
    new_name = st.text_input("नवीन व्यक्तीचे नाव टाका (Enter new person name)", placeholder="e.g. Arjun")

    if st.button("✅ Person Add करा"):
        if not new_name.strip():
            st.error("⚠️ कृपया नाव टाका!")
        elif new_name.strip().title() in st.session_state.persons:
            st.warning(f"⚠️ **{new_name.strip().title()}** आधीच list मध्ये आहे!")
        else:
            st.session_state.persons.append(new_name.strip().title())
            st.success(f"✅ **{new_name.strip().title()}** यशस्वीरित्या add केले!")
            st.rerun()

    st.markdown("---")
    st.markdown("**⚠️ Note:** नवीन add केलेले person फक्त या session पुरते आहेत. Streamlit app restart झाल्यावर default persons (Suhail, Mayur, Rahul) परत येतील. Permanent save साठी code मधील `PERSONS` list update करा.")

# ─── Footer ───────────────────────────────────────────────────────────────────
st.markdown("---")
st.markdown('<p style="text-align:center; color:#4a5568; font-size:0.8rem;">💰 Expense Tracker Portal • Built with Streamlit</p>', unsafe_allow_html=True)
